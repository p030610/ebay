from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import pandas as pd
import time
import warnings
from openpyxl import load_workbook
import os
import urllib.request
import datetime

warnings.filterwarnings("ignore")

command = input("1:크롤링 2:이미지 다운로더 3:재고확인기")

if command == '1' : 
    options = webdriver.ChromeOptions()
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    # options.add_argument("--headless")
    # driver = webdriver.Chrome('chromedriver.exe', options=options)
    driver = webdriver.Chrome('./chromedriver', options=options)
    id_pw_list = pd.read_excel("./account.xlsx")
        # Category  brand	Condition	Material	Color	country 	size	Style	정상가	할인가
    df = pd.DataFrame(columns=["ebay item number", "stock", "image", "URL","name", "Category","	Brand","Condition","Material","Color","size", "gender","bag height", "bag length","Style","Country","배송기간","배송비","정상가","할인가"])

    df.to_excel("./result.xlsx")

    workbook_name = 'result.xlsx'
    wb = load_workbook(workbook_name)
    page = wb.active

    for name, row in id_pw_list.iterrows() :

        #자동로그인 시작

        id = row[0]
        pw = row[1]

        driver.get("https://signin.ebay.com/ws/eBayISAPI.dll?SignIn&ru=https%3A%2F%2Fwww.ebay.com%2F")

        auto = input("캡차 인증을 하셨으면 1 + 엔터. 캡차 과정을 거치지 않아도 1 + 엔터를 눌러주세요")

        time.sleep(1)

        driver.find_element_by_id("userid").send_keys(id)

        time.sleep(1)

        driver.find_element_by_id("signin-continue-btn").click()

        time.sleep(1)

        auto = input("캡차 인증을 하셨으면 1 + 엔터. 캡차 과정을 거치지 않아도 1 + 엔터를 눌러주세요")

        driver.find_element_by_id("pass").send_keys(pw)

        time.sleep(1)

        driver.find_element_by_id("sgnBt").click()

        #캡차 로그인

        auto = input("캡차 인증을 하셨으면 1 + 엔터. 캡차 과정을 거치지 않아도 1 + 엔터를 눌러주세요")

        if auto == "1" :
        
            #자동로그인 끝

            driver.get("https://www.ebay.com/mye/myebay/watchlist")

            time.sleep(1)

            total_search = driver.find_element_by_class_name("filter-link__cell").text

            total_search = total_search.replace("All Categories (","").replace(")\n- Selected","")
 
            total_search = int(total_search)

            #총 검색 개수 산출 완료

            total_page = total_search / 10

            for i in range(1,int(total_page)) :
                print("총 " + str(int(total_page)) + "페이지중" + str(i) + "페이지 크롤링 중")
                driver.get("https://www.ebay.com/mye/myebay/watchlist?custom_list_id=WATCH_LIST&page=" + str(i))

                elems = driver.find_elements_by_class_name("title")
                
                links = []

                for elem in elems : 
                    if elem.get_attribute("href") != None :
                        links.append(elem.get_attribute("href"))

                for link in links : 

                    print(str(link))

                    brand = ""
                    material = ""
                    color = ""
                    country = ""
                    size = ""
                    style = ""
                    img = "이미지 개수"
                    name="상품명"
                    delivery_date = ""
                    stock = "재고 있음"
                    bag_height=""
                    bag_length=""

                    url = link
                    #url 저장
                    ebay_item_number = str(link).replace("https://www.ebay.com/itm/","")
                    
                    try : 
                        driver.get(str(link))
                    except : 
                        continue

                    time.sleep(1)

                    #상품명 찾기
                    #이미지 개수 찾기

                    #페이지 재고 소진되었는지 확인하는 로직

                    try : 
                        in_stock = driver.find_element_by_class_name("msgTextAlign").text 
                        if "ended" in in_stock : 
                            #재고 소진
                            stock = "재고 없음"
                    except NoSuchElementException :
                        pass

                    #item specifics 찾기
                    try :
                        item_specifics = driver.find_element_by_xpath('//*[@id="viTabs_0_is"]/div/div[3]/div').text
                    except NoSuchElementException : 
                        item_specifics = driver.find_element_by_xpath('//*[@id="viTabs_0_is"]/div/div[2]/div').text

                    #2번째 div가 pre-owned라고 하면서 총 div 가 2개,3개가 되는 경우로 나뉨

                    item_specific_list = []

                    for j in item_specifics.split("\n") :
                        item_specific_list.append(j)
                    
                    try : 
                        discount_price = driver.find_element_by_class_name("mainPrice").text
                    except : 
                        discount_price = "Bidding"

                    try : 
                        price = driver.find_element_by_class_name("discountPrice").find_element_by_class_name("vi-originalPrice").text.replace("Was:","")
                    except :
                        price = ""

                    if price == "" and discount_price != 'bidding' : 
                        price = discount_price
                        discount_price = ""

                    try : 
                        name = driver.find_element_by_class_name("x-item-title__mainTitle").text
                    except : 
                        continue

                    img = len(driver.find_elements_by_class_name("v-pic-item"))

                    try : 
                        country = driver.find_element_by_class_name("iti-eu-bld-gry").text
                        delivery_date = driver.find_element_by_class_name("vi-acc-del-range").text
                    except NoSuchElementException : 
                        country = ""
                        delivery_date = ""

                    try :
                        category = driver.find_element_by_class_name('vi-VR-brumblnkLst').text
                    except : 
                        category = driver.find_element_by_class_name('vim-breadcrumb').text
                    category = category.split("\n").pop()

                    condition = driver.find_element_by_class_name('d-item-condition-text').text

                    try :
                        delivery_price = driver.find_element_by_id('fshippingCost').text
                    except :
                        delivery_price = ""

                    for k in range(0, len(item_specific_list) - 1) : 

                        # print(item_specific_list[k])
                        # print(type(item_specific_list[k]))

                        if "Brand" in item_specific_list[k] :
                            brand = item_specific_list[k + 1]

                        if "Material" in item_specific_list[k] :
                            material = item_specific_list[k + 1]

                        if "Color" in item_specific_list[k] :
                            color = item_specific_list[k + 1]

                        if "Size" in item_specific_list[k] and ("Type" in item_specific_list[k]) == False:
                            size = item_specific_list[k + 1]

                        if "Style" in item_specific_list[k] :
                            style = item_specific_list[k + 1]
                            
                        if "Bag Height" in item_specific_list[k] : 
                            bag_height = item_specific_list[k + 1]

                        if "Bag Length" in item_specific_list[k] : 
                            bag_length = item_specific_list[k + 1]

                        if "Gender" in item_specific_list[k] : 
                            gender = item_specific_list[k + 1]

                    format = '%Y.%a.%b.%d'
                    format_day = "%d"
                    #Wed. Apr. 27 and Thu. May. 12
                    delivery_list = delivery_date.replace(" ","").split("and")

                    today = datetime.datetime.today()

                    try : 
                        delivery_start = datetime.datetime.strptime("2022." + delivery_list[0],format)
                        delivery_start = delivery_start - today
                        delivery_end = datetime.datetime.strptime("2022." + delivery_list[1],format)    
                        delivery_end = delivery_end - today

                        delivery  = str(delivery_start.days) + "일-" + str(delivery_end.days) + "일"
                    except : 
                        delivery = ""

                    page.append(["",ebay_item_number, stock, img, url, name, category,brand,condition,material,color,size,gender, bag_height, bag_length, style,country,delivery, delivery_price, price,discount_price])
                
                    wb.save(filename=workbook_name)

    #정상가 : 할인되기 전 가격
    #할인가 : 할인 후 가격
    # "ebay item number", "stock", "image", "URL","name", "Category","	Brand","Condition","Material","Color","size","bag height", "bag length","Style","Country","배송기간","정상가","할인가"

elif command == '2' : 
    #이미지 다운로더
    df_img = pd.read_excel("./image_input.xlsx")

    options = webdriver.ChromeOptions()
    options.add_argument("--window-size=1920,1080")
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    # options.add_argument("--headless")

    # driver = webdriver.Chrome('chromedriver.exe', options=options)
    driver = webdriver.Chrome('./chromedriver', options=options)
    

    for name, row in  df_img.iterrows() :
        try : 

            driver.get(row[0])

            time.sleep(0.5)

            os.mkdir("./images/" + str(row[1]))

            driver.find_element_by_class_name("vi-img-overlay--trans").click()

            time.sleep(1)

            # input("pause")

            elems = driver.find_element_by_id("viEnlargeImgLayer_layer_fs").find_elements_by_tag_name("img")

            index = 1

            for i in elems :
                
                url = i.get_attribute("src")

                url = url.replace("l64","l1600")

                print("링크 : " + str(row[0]) + " " + str(index) + "번 이미지 다운로드중")

                if index != 1 : 
                    urllib.request.urlretrieve(url, "./images/" + str(row[1]) + "/" + str(row[1]) + "_" + str(index) + ".jpg")

                else : 
                    urllib.request.urlretrieve(url, "./대표사진/" + str(row[1]) + "_" + str(index) + ".jpg")

                index += 1
        except : 
            pass

    print("다운로드 종료")

elif command == "3" : 
    df_stock = pd.read_excel("./stock_input.xlsx")
    options = webdriver.ChromeOptions()
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    # options.add_argument("--headless")
    # driver = webdriver.Chrome('chromedriver.exe', options=options)
    driver = webdriver.Chrome('./chromedriver', options=options)

    df_stock_output = pd.DataFrame(columns=["url","ebay item number", "재고 유무"])

    df_stock_output.to_excel("./stock_output.xlsx")

    workbook_name = 'stock_output.xlsx'
    wb = load_workbook(workbook_name)
    page = wb.active

    for name, row in df_stock.iterrows() :
        driver.get(row[0])

        time.sleep(1)
        
        try : 
            message = driver.find_element_by_class_name("statusContent").text
        except : 
            message = ""

        if "end" in message or "품절" in message: 
            page.append(["",row[0],row[0].replace("https://www.ebay.com/itm/",""),"X"])
        else : 
            page.append(["",row[0],row[0].replace("https://www.ebay.com/itm/",""),"O"])
        wb.save("./stock_output.xlsx")

    print("재고확인 종료")